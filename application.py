# coding=utf-8

import datetime
import os
from functools import wraps

import flask_excel as excel
from flask import Flask, session, render_template, request, flash, redirect, url_for, send_from_directory
from flask_babel import Babel, gettext, lazy_gettext
from flask_bcrypt import Bcrypt
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_mail import Mail, Message
from flask_wtf import FlaskForm
from openpyxl import Workbook
from openpyxl.compat import range
from wtforms import RadioField, TextAreaField, StringField, SubmitField, SelectField, PasswordField
from wtforms.validators import Email, Length, DataRequired, InputRequired

import db
from mail_settings import config_email

_ = gettext

app = Flask(__name__)
config_email(app)
app.config['SECRET_KEY'] = 'Super Secret Unguessable Key'
mail = Mail(app)

bcrypt = Bcrypt(app)
login_mgr = LoginManager(app)

excel.init_excel(app)

babel = Babel(app)
app.config['BABEL_DEFAULT_LOCALE'] = 'es'


@babel.localeselector
def get_locale():
    return request.accept_languages.best_match(['es', 'en'])


@app.before_request
def before():
    if not app.testing:
        db.open_db_connection()


# noinspection PyUnusedLocal
@app.teardown_request
def after(exception):
    if not app.testing:
        db.close_db_connection()


# INDEX + MAP + Dashboard ##############################################


@app.route('/')
def index():
    """Index page: map of all home groups"""
    return render_template('index.html')


@app.route('/downloads/<path:filename>')
def download_file(filename):
    return send_from_directory(os.getcwd(), filename, as_attachment=True)


@app.route('/dashboard')
def dashboard():
    """Display the dashboard depending on user role."""
    email = current_user.email
    role = current_user.role
    if role == 'homegroup_leader':
        homegroup_id = db.find_user_homegroup(email)
        return redirect(url_for('homegroup', homegroup_id=homegroup_id))
    if role == "admin":
        return redirect(url_for('admin_home'))
    return redirect(url_for('index'))


@app.route('/map')
def homegroup_map():
    """Display the map of all the homegroups."""
    homegroups = db.get_all_homegroup_info()
    return render_template('map.html', homegroups=homegroups)


# Display the FAQ page
@app.route('/faq')
def faq():
    locale = get_locale() or 'es'
    faq_file_name = 'faq_{}.html'.format(locale)
    return render_template(faq_file_name)


# Display the homegroup leader FAQ page
@app.route('/faq/homegroup_leader')
def faq_homegroup_leader():
    return render_template('faq_homegroup_leader.html')


# Display the admin faq page
@app.route('/faq/admin')
def faq_admin():
    return render_template('faq_admin.html')


class ContactForm(FlaskForm):
    name = StringField(lazy_gettext('Name'), validators=[DataRequired()])
    email = StringField(lazy_gettext('Email Address'), validators=[DataRequired()])
    message = TextAreaField(lazy_gettext('Message'), validators=[DataRequired()])
    submit = SubmitField(lazy_gettext('Send Email'))


# Display the contact page
@app.route('/contact', methods=['GET', 'POST'])
def contact():
    recipient_list = []
    contact_form = ContactForm()
    if contact_form.validate_on_submit():
        name = contact_form.name.data
        email = contact_form.email.data
        recipient_list.append('verbovelocity@gmail.com')  # FIXME: Keep using this address?
        message = contact_form.message.data
        email_html = render_template('contact_email.html', name=name, email=email, message=message)
        msg = Message(
            lazy_gettext('Message Received'),
            sender=email,
            recipients=recipient_list,
            html=email_html)
        mail.send(msg)
        flash(lazy_gettext('Email Sent!'), category="success")
        return redirect(url_for('index'))
    return render_template('contact.html', form=contact_form)


# USER + LOGIN + Profile/Settings ##############################################


def requires_roles(*roles):
    """Allow/disallow users from accessing pages based on their roles"""

    def wrapper(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if not hasattr(current_user, 'role'):
                flash(lazy_gettext('User does not have sufficient privileges'), category="danger")
                return redirect(url_for('index'))
            elif current_user.role not in roles:
                flash(lazy_gettext('User does not have sufficient privileges'), category="danger")
                return redirect(url_for('index'))
            return f(*args, **kwargs)

        return wrapped

    return wrapper


class UserForm(FlaskForm):
    password = PasswordField(lazy_gettext('Temporary Password'), validators=[DataRequired()])
    role = SelectField(lazy_gettext('Change Role'), choices=[], coerce=int)
    homegroups = SelectField(lazy_gettext('Choose Homegroup'), choices=[], coerce=int)
    submit = SubmitField(lazy_gettext('Create User'))


class RoleForm(FlaskForm):
    role = SelectField(lazy_gettext('Change Role'), choices=[], coerce=int)
    homegroups = SelectField(lazy_gettext('Choose Homegroup'), choices=[], coerce=int)
    submit = SubmitField(lazy_gettext('New Role'))


@app.route('/user/create/<member_id>', methods=['GET', 'POST'])
def create_user(member_id):
    """Creates a new user and hashes their password in the database."""
    email_list = []
    member = db.find_member(member_id)
    email = member['email']
    user_form = UserForm()
    user_form.role.choices = make_role_list()
    homegroups = db.get_all_homegroups()
    homegroup_list = []
    for homegroup in homegroups:
        homegroup_list.append((homegroup['id'], homegroup['name']))
    user_form.homegroups.choices = homegroup_list
    if request.method == "POST":
        email_list.append(email)
        password = user_form.password.data
        pw_hash = bcrypt.generate_password_hash(password).decode('utf-8')
        if db.has_active_role(member_id):
            flash(lazy_gettext("User already has account"), category="danger")
            return redirect(url_for('get_roles'))
        db.create_member_role(member_id, pw_hash, user_form.role.data)
        user = db.find_user(email)
        email_html = render_template('user_account_email.html', email=email, password=password, user_id=user['id'])
        msg = Message(
            lazy_gettext('User account created for Verbo Velocity'),
            sender='verbovelocity@gmail.com',  # FIXME: Appears multiple times
            recipients=email_list,
            html=email_html)
        mail.send(msg)
        if user_form.role.data == 1:
            home_group_id = user_form.homegroups.data
            db.add_leader_to_homegroup(member_id, home_group_id)

        flash(lazy_gettext('User Created'), category="success")
        return redirect(url_for('get_roles'))
    return render_template('create_user.html', form=user_form, email=email)


@app.route('/roles')
@login_required
@requires_roles('admin')
def get_roles():
    """Show role page."""
    member_roles = db.get_all_member_roles()
    return render_template('roles.html', member_roles=member_roles)


def make_role_list():
    role_list = []
    all_roles = db.get_all_roles()
    for role in all_roles:
        # FIXME: Massive translation hack; need to get localized version from DB
        label = 'Líder de Iglesia de Hogar'
        if role["role"] == 'admin':
            label = 'Administrador'
        role_list.append((role["id"], label))
    return role_list


@app.route('/roles/new_role/<member_id>', methods=['GET', 'POST'])
def assign_new_role(member_id):
    """Create a new role for the user that already exists."""
    current_user = db.find_user(session['username'])['member_id']
    if str(member_id) == str(current_user):
        flash(lazy_gettext("You cannot edit your own role - please contact a system administrator"), category="warning")
        return redirect(url_for('get_roles'))
    member = db.find_member(member_id)
    email = member['email']
    user_form = RoleForm()
    user_form.role.choices = make_role_list()
    homegroups = db.get_all_homegroups()
    homegroup_list = []
    for homegroup in homegroups:
        homegroup_list.append((homegroup['id'], homegroup['name']))
    user_form.homegroups.choices = homegroup_list

    if request.method == "POST":

        db.assign_new_role(member_id, user_form.role.data)
        if user_form.role.data == 1:
            homegroup_id = user_form.homegroups.data
            db.add_leader_to_homegroup(member_id, homegroup_id)

        flash(lazy_gettext('Role Created'), category="success")
        return redirect(url_for('get_roles'))
    return render_template('assign_role.html', form=user_form, email=email)


@app.route('/roles/edit/<member_id>/<role_id>')
@login_required
@requires_roles('admin')
def edit_role(member_id, role_id):
    current_user = db.find_user(session['username'])['member_id']
    if str(member_id) == str(current_user):
        flash(lazy_gettext("You cannot edit your own role - please contact a system administrator"), category="warning")
        return redirect(url_for('get_roles'))
    is_active = db.is_role_active(member_id, role_id)
    active = '1'
    if is_active:
        active = '0'
    db.update_role(member_id, role_id, active)
    return redirect(url_for('get_roles'))


@app.route('/hgleader/<member_id>/<homegroup_id>')
@login_required
@requires_roles('admin')
def deactivate_hgleader(member_id, homegroup_id):
    db.deactivate_hgleader(member_id, homegroup_id)
    return redirect(url_for('get_roles'))


class UpdateUserForm(FlaskForm):
    """Update password form"""
    old_password = PasswordField(lazy_gettext('Current Password'), validators=[DataRequired()])
    new_password = PasswordField(lazy_gettext('New Password'), validators=[DataRequired()])
    confirm_password = PasswordField(lazy_gettext('Confirm Password'), validators=[DataRequired()])
    submit = SubmitField(lazy_gettext('Update Password'))


@app.route('/user/edit/<user_id>', methods=['GET', 'POST'])
def update_user(user_id):
    """Update password."""
    member = db.find_user_info(user_id)
    email = member['email']
    user_form = UpdateUserForm(email=member['email'])
    if user_form.validate_on_submit():
        old_password = user_form.old_password.data
        new_password = user_form.new_password.data
        confirm_password = user_form.confirm_password.data
        if bcrypt.check_password_hash(member['password'], old_password):
            if new_password == confirm_password:
                password = new_password
                pw_hash = bcrypt.generate_password_hash(password).decode('utf-8')
                db.update_password(user_id, pw_hash, member['role_id'])
                flash(lazy_gettext('Password updated'), category="success")
                return redirect(url_for('index'))
            else:
                flash(lazy_gettext('New password and confirmation password do not match'), category="danger")
                return redirect(url_for('update_user', user_id=user_id))
        else:
            flash(lazy_gettext('Entered wrong current password'), category="danger")
            return redirect(url_for('update_user', user_id=user_id))
    return render_template('update_user.html', form=user_form, email=email)


class User(object):
    """Class for the currently logged-in user (if there is one). Only stores the user's email."""

    def __init__(self, email):
        self.email = email
        user = db.find_user(self.email)
        if user is not None:
            self.role = user['role']
            self.name = db.find_member_info(self.email)['first_name']
            self.user_id = user['id']
            self.member_id = db.find_member_info(self.email)['id']
        else:
            self.role = 'no role'
            self.name = 'no name'
        if self.role == 'homegroup_leader':
            self.homegroup_id = db.find_user_homegroup(self.email)

        self.is_authenticated = True
        self.is_active = True
        self.is_anonymous = False

    def get_id(self):
        """Return the unique ID for this user. Used by Flask-Login to keep track of the user in the session object."""
        return self.email

    def get_role(self):
        """Returns the role of the user from the db"""
        return self.role

    def __repr__(self):
        return "<User '{}' {} {} {} {}>".format(self.email, self.role, self.is_authenticated, self.is_active,
                                                self.is_anonymous)


@login_mgr.user_loader
def load_user(user_id):
    """Return the currently logged-in user when given the user's unique ID"""
    return User(user_id)


# checks to see if the password entered matches the hash password
def authenticate(email, password):
    """Check whether the arguments match a user from the "database" of valid users."""
    valid_users = db.get_all_users()

    for user in valid_users:
        if email == user['email'] and bcrypt.check_password_hash(user['password'], password):
            return email
    return None


class LoginForm(FlaskForm):
    email = StringField(lazy_gettext('Email Address'), validators=[DataRequired()])
    password = PasswordField(lazy_gettext('Password'), validators=[DataRequired()])
    submit = SubmitField(lazy_gettext('Log In'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Log the user in, create a new session, and create a current user of class 'User'"""
    login_form = LoginForm()

    if login_form.validate_on_submit() and login_form.validate():
        user = db.find_user(login_form.email.data)
        if user:
            member_id = user['member_id']
            is_active = db.has_active_role(member_id)
        else:
            is_active = False

        if authenticate(login_form.email.data, login_form.password.data) and is_active:
            # Credentials authenticated.
            # Create the user object, let Flask-Login know, and redirect to the home page
            current_user = User(login_form.email.data)
            login_user(current_user)
            session['username'] = current_user.email
            return redirect(url_for('dashboard'))
        else:
            # Authentication failed.
            flash(lazy_gettext('Invalid email address or password'), category="danger")
            return redirect(url_for('login'))

    return render_template('login.html', form=login_form)


@app.route('/logout')
def logout():
    """Log the user out and removes the session."""
    logout_user()
    return redirect(url_for('index'))


@app.route('/user/profile/<user_id>')
@login_required
def user_profile(user_id):
    user_info = db.find_user_info(user_id)
    member = db.find_member_info(user_info['email'])
    return redirect(url_for('edit_member', member_id=member['id']))


# HOME GROUP  (Home Group Leader)##############################################


@app.route('/homegroup/<homegroup_id>')
@login_required
@requires_roles('homegroup_leader', 'admin')
def homegroup(homegroup_id):
    homegroup = db.find_homegroup_by_id(homegroup_id)
    attendance_count = db.get_homegroup_attendance_counts(homegroup_id)
    count_members = db.number_of_members_in_homegroup(homegroup_id)
    my_date = datetime.datetime.now()
    month_string = my_date.strftime("%B").upper()
    now = datetime.datetime.now()
    month = now.month
    hg_attendance_rate = str(int(db.get_homegroup_attendance_rate(month, homegroup_id))) + '%'
    number_meetings = db.number_of_meetings_held(homegroup_id)
    if not attendance_count:
        if current_user.role == "admin":
            flash(lazy_gettext("No attendance data for this Home Group"), category="warning")
            return redirect(url_for('get_homegroups', countMembers=count_members, homegroup_id=homegroup_id))
        else:
            return render_template('homegroup.html', countMembers=count_members, currentHomegroup=homegroup,
                                   numMeetings=number_meetings,
                                   attendance_count=attendance_count, member_attendance=[], dates=[])
    # member_attendance = db.homegroup_member_attendance(homegroup_id)
    members = db.get_homegroup_members(homegroup_id)
    member_attendance = []
    dates = db.get_last_3_dates(homegroup_id)
    for member in members:
        attendance = db.get_member_attendance(homegroup_id, member['member_id'])
        names = []
        if attendance:
            name = attendance[0]['first_name'] + ' ' + attendance[0]['last_name']
            names.append(name)
            for date in dates:
                has_date = False
                for item in attendance:
                    if date['date'] == item['date'] and item['attendance']:
                        names.append(item['attendance'])
                        has_date = True
                if not has_date:
                    names.append(False)
                list_length = len(names)
            if list_length < 4:
                for _ in range(0, (4 - list_length)):
                    names.append(False)
            member_attendance.append(names)

        else:
            names.append(member['first_name'] + " " + member['last_name'])
            for _ in range(0, 4):
                names.append(False)
            member_attendance.append(names)

    return render_template('homegroup.html', numMeetings=number_meetings, attendance_rate=hg_attendance_rate,
                           currentMonth=month_string, countMembers=count_members, currentHomegroup=homegroup,
                           attendance_count=attendance_count, member_attendance=member_attendance, dates=dates)


# this is the homegroup main page / dashboard
@app.route('/homegroup_split/<homegroup_id>', methods=['GET', 'POST'])
@login_required
@requires_roles('admin')
def homegroup_split(homegroup_id):
    new_homegroup = CreateHomeGroupSplitForm()

    if request.method == "POST" and new_homegroup.validate():
        name = new_homegroup.name.data
        location = new_homegroup.location.data
        description = new_homegroup.description.data
        latitude = new_homegroup.latitude.data
        longitude = new_homegroup.longitude.data

        group1 = db.create_homegroup(name, location, description, latitude, longitude)
        homegroup1 = group1['id']
        name2 = new_homegroup.name2.data
        location2 = new_homegroup.location2.data
        description2 = new_homegroup.description2.data
        latitude2 = new_homegroup.latitude2.data
        longitude2 = new_homegroup.longitude2.data
        group2 = db.create_homegroup(name2, location2, description2, latitude2, longitude2)
        homegroup2 = group2['id']
        if group1['rowcount'] == 1 and group2['rowcount'] == 1:
            return redirect(url_for('homegroup_split_members', homegroup_id=homegroup_id, homegroup_id1=homegroup1,
                                    homegroup_id2=homegroup2))

    return render_template('homegroup_split.html', form=new_homegroup)


# this is the homegroup main page / dashboard
@app.route('/split_member/<member_id>/<homegroup_id>/<homegroup_id1>/<homegroup_id2>/<group>', methods=['GET', 'POST'])
@login_required
@requires_roles('admin')
def split_member(member_id, homegroup_id, homegroup_id1, homegroup_id2, group):
    db.remove_member(homegroup_id, member_id)
    if group == 'one':
        db.remove_member(homegroup_id2, member_id)
        db.add_member_to_homegroup(homegroup_id1, member_id)
    if group == 'two':
        db.remove_member(homegroup_id1, member_id)
        db.add_member_to_homegroup(homegroup_id2, member_id)
    return redirect(url_for('homegroup_split_members', homegroup_id=homegroup_id, homegroup_id1=homegroup_id1,
                            homegroup_id2=homegroup_id2))


@app.route('/homegroup_split_members/<homegroup_id>/<homegroup_id1>/<homegroup_id2>', methods=['GET', 'POST'])
@login_required
@requires_roles('admin')
def homegroup_split_members(homegroup_id, homegroup_id1, homegroup_id2):
    """Add members to split home group."""
    members = db.get_split_homegroup_info(homegroup_id, homegroup_id1, homegroup_id2)
    current_homegroup = homegroup_id
    homegroup_to_be_split = db.find_homegroup_by_id(homegroup_id)
    homegroup1 = db.find_homegroup_by_id(homegroup_id1)
    homegroup2 = db.find_homegroup_by_id(homegroup_id2)
    homegroup_to_be_split_count = 0
    homegroup1_count = 0
    homegroup2_count = 0
    for member in members:
        if member['group'] == homegroup1['id']:
            homegroup1_count += 1
        elif member['group'] == homegroup2['id']:
            homegroup2_count += 1
        else:
            homegroup_to_be_split_count += 1

    return render_template('homegroup_split_members.html', homegroup_to_be_split=homegroup_to_be_split,
                           homegroup1=homegroup1, homegroup2=homegroup2, currentHomegroup=current_homegroup,
                           old_members=members, homegroup1_count=homegroup1_count, homegroup2_count=homegroup2_count,
                           homegroup_to_be_split_count=homegroup_to_be_split_count)


@app.route('/homegroup_analytics')
@login_required
@requires_roles('admin')
def homegroup_analytics():
    """Return further analytics for homegroups."""
    year = datetime.datetime.now().strftime("%Y")
    analytics = db.homegroup_analytics(year)
    minors = db.number_of_minors(year)
    new_members = db.number_of_new_members(year)
    attending_members = db.members_attending_a_homegroup(year)
    total_members = db.total_members(year)
    analytics_list = ['', '', '', '', '', '', '', '', '', '', '', '']
    minors_list = ['', '', '', '', '', '', '', '', '', '', '', '']
    new_members_list = ['', '', '', '', '', '', '', '', '', '', '', '']
    attending_members_list = ['', '', '', '', '', '', '', '', '', '', '', '']
    total_members_list = ['', '', '', '', '', '', '', '', '', '', '', '']
    for row in analytics:
        index = int(row['month']) - 1
        analytics_list[index] = int(row['count'])
    for row in minors:
        index = int(row['month']) - 1
        minors_list[index] = int(row['count'])
    for row in new_members:
        index = int(row['month']) - 1
        new_members_list[index] = int(row['count'])
    for row in attending_members:
        index = int(row['month']) - 1
        attending_members_list[index] = int(row['count'])
    for row in total_members:
        index = int(row['month']) - 1
        total_members_list[index] = int(row['count'])

    return render_template('homegroup_analytics.html', totalMembersList=total_members_list,
                           attendingMembersList=attending_members_list, year=year, newMembersList=new_members_list,
                           minorsList=minors_list, analytics=analytics_list)


class AttendanceForm(FlaskForm):
    # member_id = StringField('member Id', validators=[Length(min=1, max=40)])
    # meeting_id = StringField('Meeting Id', validators=[Length(min=1, max=40)])
    radio = RadioField(lazy_gettext('Attendance'), choices=["y", "n"])
    submit = SubmitField(lazy_gettext('Submit'))


# this is the default attendance page (allows you to select date/time then generate an attendance report)
@app.route('/homegroup/attendance/<homegroup_id>', methods=['GET', 'POST'])
@login_required
@requires_roles('homegroup_leader')
def attendance(homegroup_id):
    attendance_form = AttendanceForm()
    members = db.get_homegroup_members(homegroup_id)
    show_members = 'N'
    if request.method == "POST":
        date = request.form['AttendanceDate']
        time = request.form['AttendanceTime']
        meeting_id = db.add_date(date, time)['id']

        db.generate_attendance_report(homegroup_id, meeting_id)
        return redirect(url_for('edit_attendance', homegroup_id=homegroup_id, meeting_id=meeting_id))
    return render_template('attendance.html', currentHomegroup=homegroup_id, form=attendance_form, members=members,
                           showmembers=show_members)


# Sends email if user and missed a certain number of meetings
def system_notify_member(member_id, num_misses):
    homegroup_id = db.find_member_homegroup(member_id)['homegroup_id']
    leader = db.find_homegroup_leader(homegroup_id)
    email = db.find_member(member_id)['email']
    email_list = [email]
    leader_email = leader['email']
    leader_phone = leader['phone_number']
    leader_name = leader['first_name'] + ' ' + leader['last_name']
    email_html = render_template('notify_member_email.html', num_misses=num_misses, leader_name=leader_name,
                                 leader_phone=leader_phone, leader_email=leader_email)
    msg = Message(
        lazy_gettext('System Reminder: Missing meetings'),
        sender='verbovelocity@gmail.com',
        recipients=email_list,
        html=email_html)
    mail.send(msg)


@app.route('/homegroup/data/<homegroup_id>', methods=['GET', 'POST'])
@login_required
def homegroup_data(homegroup_id):
    attendance = db.get_homegroup_attendance_records(homegroup_id)
    hg_name = db.find_homegroup_by_id(homegroup_id)['name']
    wb = Workbook()
    dest_filename = hg_name + '-' + gettext('attendance') + '.xlsx'
    ws1 = wb.active
    ws1.title = gettext("Attendance")
    row_num = 2
    for row in attendance:
        col_num = 1
        for col in row:
            if row_num == 2:
                _ = ws1.cell(column=col_num, row=1, value=col)
            _ = ws1.cell(column=col_num, row=row_num, value=row[col])
            col_num = col_num + 1
        row_num = row_num + 1
    wb.save(filename=dest_filename)
    return redirect(url_for('download_file', filename=dest_filename))


@app.route('/homegroup/data/all', methods=['GET', 'POST'])
@login_required
@requires_roles('admin')
def all_homegroup_data():
    attendance = db.get_all_homegroup_attendance_records()
    hg_name = "all-homegroup"
    wb = Workbook()
    dest_filename = hg_name + '-' + gettext('attendance') + '.xlsx'
    ws1 = wb.active
    ws1.title = gettext("Attendance")
    row_num = 2
    for row in attendance:
        col_num = 1
        for col in row:
            if row_num == 2:
                _ = ws1.cell(column=col_num, row=1, value=col)
            _ = ws1.cell(column=col_num, row=row_num, value=row[col])
            col_num = col_num + 1
        row_num = row_num + 1
    wb.save(filename=dest_filename)
    return redirect(url_for('download_file', filename=dest_filename))


# adds (or updates) a new entry of attendance into the db
def update_attendance(homegroup_id, member_id, meeting_id, attendance):
    db.update_attendance(homegroup_id, member_id, meeting_id, attendance)
    return redirect(url_for('edit_attendance', homegroup_id=homegroup_id, meeting_id=meeting_id))


class EditAttendanceForm(FlaskForm):
    submit = SubmitField(lazy_gettext('Save'))


# This allows you to edit homegroup attendance
@app.route('/homegroup/attendance/edit/<homegroup_id>/<meeting_id>', methods=['GET', 'POST'])
@login_required
@requires_roles('homegroup_leader', 'admin')
def edit_attendance(homegroup_id, meeting_id):
    att_form = EditAttendanceForm()
    members_in_attendance = db.get_attendance(homegroup_id, meeting_id)
    date = db.find_date(meeting_id)['date']
    time = db.find_date(meeting_id)['time']
    edit_or_new = 'new'
    for member in members_in_attendance:
        if member['attendance'] == 1:
            edit_or_new = 'edit'
    if att_form.validate_on_submit():
        for member in members_in_attendance:
            input_name = 'member_' + str(member['member_id'])
            if input_name in request.form:
                update_attendance(homegroup_id, member['member_id'], meeting_id, '1')
            else:
                update_attendance(homegroup_id, member['member_id'], meeting_id, '0')
                if edit_or_new == 'new':
                    attendance_dates = db.system_attendance_alert(homegroup_id, member['id'], 3)
                    for date in attendance_dates:
                        if date['attendance'] == 1:
                            pass
                    if len(attendance_dates) < 3:
                        pass

        return redirect(url_for('get_attendance_dates', homegroup_id=homegroup_id))

    return render_template('edit_attendance.html', currentHomegroup=homegroup_id, meeting_id=meeting_id,
                           members=members_in_attendance, date=date, time=time, form=att_form)


# Return all the attendance dates -- this is for the attendance reports page
@app.route('/homegroup/attendance/dates/<homegroup_id>', methods=['GET'])
@login_required
@requires_roles('homegroup_leader', 'admin')
def get_attendance_dates(homegroup_id):
    return render_template('attendance_reports.html', currentHomegroup=homegroup_id,
                           records=db.get_attendance_dates(homegroup_id))


# View attendance history for a particular
@app.route('/homegroup/attendance/view/<homegroup_id>', methods=['GET'])
@login_required
@requires_roles('admin')
def view_attendance(homegroup_id):
    homegroup = db.find_homegroup_by_id(homegroup_id)
    attendance_count = db.get_homegroup_attendance_counts(homegroup_id)
    return render_template('view_attendance.html', currentHomegroup=homegroup,
                           attendance_count=attendance_count, myHomegroup=homegroup_id,
                           records=db.get_attendance_dates(homegroup_id))


@app.route('/homegroup/attendance/view/report/<homegroup_id>/<meeting_id>', methods=['GET'])
@login_required
@requires_roles('homegroup_leader', 'admin')
def view_attendance_report(homegroup_id, meeting_id):
    members_in_attendance = db.get_attendance(homegroup_id, meeting_id)
    attendance_count = db.get_homegroup_attendance_counts(homegroup_id)
    date = db.find_date(meeting_id)['date']
    time = db.find_date(meeting_id)['time']
    return render_template('view_attendance_report.html', currentHomegroup=homegroup_id, meeting_id=meeting_id,
                           members=members_in_attendance, date=date, time=time, attendance_count=attendance_count)


# Edit a particular homegroup
@app.route('/homegroup/edit/<homegroup_id>', methods=['GET', 'POST'])
@login_required
@requires_roles('homegroup_leader', 'admin')
def edit_homegroup(homegroup_id):
    row = db.find_homegroup_by_id(homegroup_id)
    hg_form = CreateHomeGroupForm(name=row['name'],
                                  description=row['description'],
                                  location=row['location'],
                                  latitude=row['latitude'],
                                  longitude=row['longitude'])
    if request.method == "POST" and hg_form.validate():
        name = hg_form.name.data
        description = hg_form.description.data
        location = hg_form.location.data
        latitude = hg_form.latitude.data
        longitude = hg_form.longitude.data
        rowcount = db.edit_homegroup(homegroup_id, name, location, description, latitude, longitude)
        if rowcount == 1:
            flash(lazy_gettext("Home Group updated"), category="success")
            if current_user.role == 'admin':
                return redirect(url_for('get_homegroups'))
            return redirect(url_for('homegroup', homegroup_id=homegroup_id))

    return render_template('edit_homegroup.html', form=hg_form)


# this is the iframe that is in the creating/editing homegroup -- allows you to type in address and finds location
@app.route('/homegroup/select_location')
def select_location():
    return render_template('select_location.html')


# this is the iframe that is in the creating/editing homegroup -- allows you to type in address and finds location
@app.route('/homegroup/select_location2')
def select_location2():
    return render_template('select_location2.html')


# MEMBER (Home Group Leader) ##############################################


@app.route('/homegroup/member/new/<homegroup_id>')
@login_required
@requires_roles('homegroup_leader', 'admin')
def member_search(homegroup_id):
    members = db.get_all_members_not_in_homegroup(homegroup_id)
    return render_template('member_search.html', all_members=members, homegroup_id=homegroup_id)


# Add a member to a particular homegroup
@app.route('/homegroup/member/add/<homegroup_id>/<member_id>')
@login_required
@requires_roles('homegroup_leader', 'admin')
def add_member_to_homegroup(homegroup_id, member_id):
    inactive_homegroup_members = db.get_homegroup_inactive_members(homegroup_id)
    new = 'Y'
    for members in inactive_homegroup_members:
        if int(members['member_id']) == int(member_id):
            new = 'N'
            db.reactive_homegroup_member(homegroup_id, member_id)
    if new == 'Y':
        member_home_group = db.member_already_in_homegroup(member_id)
        if member_home_group:
            leader = db.find_homegroup_leader(member_home_group['homegroup_id'])

            flash("Member is in another Home Group, contact {} to remove member".format(
                leader['first_name'] + " " + leader['last_name'] + " " + leader['phone_number']), "danger")
            members = db.get_all_members_not_in_homegroup(homegroup_id)
            return render_template('member_search.html', all_members=members, homegroup_id=homegroup_id)
        else:
            db.add_member_to_homegroup(homegroup_id, member_id)
            member = db.find_member(member_id)
            flash(lazy_gettext("Member {} added to homegroup").format(member['first_name'] + " " + member['last_name']),
                  category="success")

    return redirect(url_for('get_homegroup_members', homegroup_id=homegroup_id))


@app.route('/member/create', methods=['GET', 'POST'])
@login_required
@requires_roles('admin')
def create_member():
    """Create a member (not associated with a homegroup)"""
    return _create_new_member(url_for('all_members'))


@app.route('/homegroup/create_member/<homegroup_id>', methods=['GET', 'POST'])
@login_required
@requires_roles('homegroup_leader', 'admin')
def create_new_member_for_homegroup(homegroup_id):
    """Create a new member for a particular homegroup"""
    return _create_new_member(url_for('get_homegroup_members', homegroup_id=homegroup_id),
                              homegroup_id)


class CreateMemberForm(FlaskForm):
    first_name = StringField(lazy_gettext('First Name'),
                             [Length(min=2, max=30, message=lazy_gettext("First name is a required field"))])
    last_name = StringField(lazy_gettext('Last Name'),
                            [Length(min=2, max=30, message=lazy_gettext("Last name is a required field"))])
    email = StringField(lazy_gettext('Email'), [Email(lazy_gettext("Please enter valid email"))])
    phone_number = StringField(lazy_gettext('Phone Number'),
                               [InputRequired(message=lazy_gettext("Please enter valid phone number"))])
    gender = SelectField(lazy_gettext('Gender'), choices=[('M', lazy_gettext('Male')), ('F', lazy_gettext('Female'))])
    baptism_status = SelectField(lazy_gettext('Baptized?'),
                                 choices=[('True', lazy_gettext('Yes')), ('False', lazy_gettext('No'))])
    is_a_parent = SelectField(lazy_gettext('Do you have children?'),
                              choices=[('True', lazy_gettext('Yes')), ('False', lazy_gettext('No'))])
    how_did_you_find_out = SelectField(lazy_gettext('How did you find out about the homegroup?'), choices=[],
                                       coerce=int)
    marital_status = SelectField(lazy_gettext('Marital Status'), choices=[], coerce=int)
    submit = SubmitField(lazy_gettext('Save Member'))


def _create_new_member(redirect_url, homegroup_id=None):
    """Generic function to create a new member.

    Redirect to 'redirect_url' upon success. If 'homegroup_id' has a value,
    associate the new member with the indicated homegroup.
    """
    member_form = CreateMemberForm()

    status_list = []
    for status in db.get_all_marital_statuses():
        status_list.append((status["id"], status["marital_status_name"]))
    member_form.marital_status.choices = status_list

    method_list = []
    for method in db.get_all_how_did_you_find_out_values():
        method_list.append((method["id"], method["how_did_you_find_out_name"]))
    member_form.how_did_you_find_out.choices = method_list

    if member_form.validate_on_submit():
        if db.find_member_info(member_form.email.data):
            flash(lazy_gettext("Person with this email already exists"), category="danger")
        else:
            member_info = {
                'first_name': member_form.first_name.data,
                'last_name': member_form.last_name.data,
                'email': member_form.email.data,
                'phone_number': member_form.phone_number.data,
                'gender': member_form.gender.data,
                'birthday': request.form['Birthday'],
                'baptism_status': member_form.baptism_status.data,
                'marital_status_id': member_form.marital_status.data,
                'how_did_you_find_out_id': member_form.how_did_you_find_out.data,
                'is_a_parent': member_form.is_a_parent.data,
                'join_date': request.form['JoinDate']
            }

            member_info = db.create_member(member_info)
            if member_info['rowcount'] == 1:
                flash(lazy_gettext("Member {} {} Created").format(member_form.first_name.data,
                                                                  member_form.last_name.data),
                      category="success")
                if homegroup_id is not None:
                    db.add_member_to_homegroup(homegroup_id, member_info['id'])
                return redirect(redirect_url)

    return render_template('create_member.html', form=member_form)


@app.route('/homegroup/members/<homegroup_id>')
@login_required
@requires_roles('homegroup_leader', 'admin')
def get_homegroup_members(homegroup_id):
    """View all members in a homegroup"""
    current_homegroup = db.find_homegroup_by_id(homegroup_id)

    homegroup_members = db.get_homegroup_members(homegroup_id)
    email_addresses = []
    for member in homegroup_members:
        email_addresses.append(member["email"])
    list2 = ""
    for item in email_addresses:
        list2 = list2 + ", " + item
    return render_template('homegroup_members.html', homegroup=db.get_homegroup_members(homegroup_id),
                           currentHomegroup=current_homegroup, homegroupEmails=db.get_homegroup_emails(homegroup_id),
                           emails=list2)


# edits member information
@app.route('/member/edit/<member_id>', methods=['GET', 'POST'])
@login_required
@requires_roles('homegroup_leader', 'admin')
def edit_member(member_id):
    if int(current_user.member_id) == int(member_id):
        heading_text = lazy_gettext('Edit My Info')
    else:
        heading_text = lazy_gettext('Edit Member')

    row = db.find_member(member_id)

    member_form = CreateMemberForm(first_name=row['first_name'],
                                   last_name=row['last_name'],
                                   email=row['email'],
                                   phone_number=row['phone_number'],
                                   gender=row['gender'],
                                   baptism_status=row['baptism_status'],
                                   is_a_parent=row['is_a_parent'],
                                   how_did_you_find_out=row['how_did_you_find_out_id'],
                                   marital_status=row['marital_status_id'])
    birthday_form = row['birthday']
    join_date_form = row['join_date']

    marital_status = db.get_all_marital_statuses()
    how_did_you_find_out = db.get_all_how_did_you_find_out_values()
    status_list = []
    method_list = []
    for status in marital_status:
        status_list.append((status["id"], status["marital_status_name"]))
    member_form.marital_status.choices = status_list
    for method in how_did_you_find_out:
        method_list.append((method["id"], method["how_did_you_find_out_name"]))
    member_form.how_did_you_find_out.choices = method_list

    # TODO: add validators back!!! and member_form.validate()
    if request.method == "POST":
        first_name = member_form.first_name.data
        last_name = member_form.last_name.data
        email = member_form.email.data
        phone_number = member_form.phone_number.data
        gender = member_form.gender.data
        birthday = request.form['Birthday']
        baptism_status = member_form.baptism_status.data
        marital_status = member_form.marital_status.data
        how_did_you_find_out = member_form.how_did_you_find_out.data
        is_a_parent = member_form.is_a_parent.data
        join_date = request.form['JoinDate']

        rowcount = db.edit_member(member_id, first_name, last_name, email, phone_number, gender, birthday,
                                  baptism_status, marital_status, how_did_you_find_out, is_a_parent, join_date)
        if rowcount == 1:
            flash(lazy_gettext("Member {} Updated").format(member_form.first_name.data), category="success")
            if current_user.role == 'admin':
                return redirect(url_for('all_members'))
            else:
                homegroup_id = db.find_member_homegroup(member_id)['homegroup_id']
                return redirect(url_for('get_homegroup_members', homegroup_id=homegroup_id))

    return render_template('edit_member.html', heading_text=heading_text, form=member_form, bDay=birthday_form,
                           joinDay=join_date_form)


# removes a member from a particular homegroup
@app.route('/homegroup/member/delete/<homegroup_id>/<member_id>', methods=['GET', 'POST'])
@login_required
@requires_roles('homegroup_leader', 'admin')
def remove_member(homegroup_id, member_id):
    rowcount = db.remove_member(homegroup_id, member_id)
    if rowcount == 1:
        flash(lazy_gettext("Member Removed"), category="success")
    return redirect(url_for('get_homegroup_members', homegroup_id=homegroup_id))


# ADMIN FUNCTIONS ##############################################


# Admin - Home Group ####
class CreateHomeGroupForm(FlaskForm):
    name = StringField(lazy_gettext('Name'), [Length(min=2, max=50, message=lazy_gettext("Name is a required field"))])
    description = TextAreaField(lazy_gettext('Description'),
                                [InputRequired(message=lazy_gettext("Please enter a description"))])
    location = StringField(lazy_gettext('Address'), [InputRequired(message=lazy_gettext("Please enter valid Address"))])
    latitude = StringField(lazy_gettext('Latitude'))
    longitude = StringField(lazy_gettext('Longitude'))
    submit = SubmitField(lazy_gettext('Save Home Group'))


class CreateHomeGroupSplitForm(FlaskForm):
    name = StringField(lazy_gettext('Name'), [Length(min=2, max=50, message=lazy_gettext("Name is a required field"))])
    description = TextAreaField(lazy_gettext('Description'),
                                [InputRequired(message=lazy_gettext("Please enter a description"))])
    location = StringField(lazy_gettext('Address'), [InputRequired(message=lazy_gettext("Please enter valid Address"))])
    latitude = StringField(lazy_gettext('Latitude'))
    longitude = StringField(lazy_gettext('Longitude'))
    name2 = StringField(lazy_gettext('Name'),
                        [Length(min=2, max=50, message=lazy_gettext("Name is a required field"))])
    description2 = TextAreaField(lazy_gettext('Description'),
                                 [InputRequired(message=lazy_gettext("Please enter a description"))])
    location2 = StringField(lazy_gettext('Address'),
                            [InputRequired(message=lazy_gettext("Please enter valid Address"))])
    latitude2 = StringField(lazy_gettext('Latitude'))
    longitude2 = StringField(lazy_gettext('Longitude'))
    submit1 = SubmitField(lazy_gettext('Create Home Groups'))


# Display admin home page
@app.route('/admin')
@login_required
@requires_roles('admin')
def admin_home():
    attendance_count = db.get_attendance_counts()
    for i in range(0, len(attendance_count)):
        attendance_count[i]["month"] = lazy_gettext(attendance_count[i]["month"])
    homegroup_member_data = db.get_top_n_homegroup_member_counts('5')
    gender = db.gender_report()
    active_homegroups = db.number_of_active_homegroups()
    members = db.number_of_members_attending_homegroups()
    now = datetime.datetime.now()
    month = now.month
    attendance_rate = str(int(db.attendance_rate_for_current_month(month))) + '%'
    homegroup_leaders = db.number_of_homegroup_leaders()
    homegroups = db.number_of_homegroups()
    my_date = datetime.datetime.now()
    month_string = my_date.strftime("%B").upper()
    return render_template('admin_home.html', currentMonth=month_string, attendance_rate=attendance_rate,
                           active_homegroups=active_homegroups, members=members, homegroup_leaders=homegroup_leaders,
                           homegroups=homegroups, gender=gender, hgdata=homegroup_member_data,
                           attendance_count=attendance_count, homegroup_data=homegroup_data)


# create homegroup
@app.route('/homegroup/create', methods=['GET', 'POST'])
@login_required
@requires_roles('admin')
def create_homegroup():
    new_homegroup = CreateHomeGroupForm()

    if request.method == "POST" and new_homegroup.validate():
        name = new_homegroup.name.data
        location = new_homegroup.location.data
        description = new_homegroup.description.data
        latitude = new_homegroup.latitude.data
        longitude = new_homegroup.longitude.data
        rowcount = db.create_homegroup(name, location, description, latitude, longitude)

        if rowcount == 1:
            flash(lazy_gettext("Homegroup {} Created").format(new_homegroup.name.data), category="success")
            return redirect(url_for('get_homegroups'))

    return render_template('create_homegroup.html', form=new_homegroup)


# Show all homegroups
@app.route('/homegroup/all')
@login_required
@requires_roles('admin')
def get_homegroups():
    return render_template('homegroup_list.html', homegroup_list=db.get_all_homegroups(),
                           inactiveHomegroups=db.get_all_inactive_homegroups(), showInactive=False)


# Deactivate a homegroup
@app.route('/homegroup/delete/<homegroup_id>', methods=['GET', 'POST'])
@login_required
@requires_roles('admin')
def deactivate_homegroup(homegroup_id):
    # if the member is not active
    if not db.find_homegroup_by_id(homegroup_id)['is_active']:
        flash(lazy_gettext("Homegroup Deactivated"), category="success")
    return redirect(url_for('get_homegroups'))


# Reactivate a homegroup
@app.route('/homegroup/reactivate/<homegroup_id>', methods=['GET', 'POST'])
@login_required
@requires_roles('admin')
def reactivate_homegroup(homegroup_id):
    # if the member is not active
    if db.find_homegroup_by_id(homegroup_id)['is_active']:
        flash(lazy_gettext("Homegroup Reactivated"), category="success")
    return redirect(url_for('get_homegroups'))


# Admin - Member ###########


# shows all members
@app.route('/member/all')
@login_required
@requires_roles('admin')
def all_members():
    emails = db.get_all_members_emails()
    email_addresses = []
    for email in emails:
        email_addresses.append(email["email"])
    list2 = ""
    for item in email_addresses:
        list2 = list2 + ", " + item
    return render_template('all_members.html', members=db.get_all_members(), emails=list2,
                           inactiveMembers=db.get_all_inactive_members(), showInactive=False)


# sets a member inactive in the system
@app.route('/member/delete/<member_id>', methods=['GET', 'POST'])
@login_required
@requires_roles('admin')
def deactivate_member(member_id):
    # if the member is not active
    if db.find_member(member_id)['is_active'] == '0':
        flash(lazy_gettext("Member Deactivated"), category="success")
    return redirect(url_for('all_members'))


# sets a member active in the system
@app.route('/member/add/<member_id>', methods=['GET', 'POST'])
@login_required
@requires_roles('admin')
def reactivate_member(member_id):
    if db.find_member(member_id)['is_active'] == '1':
        flash(lazy_gettext("Member Reactivated"), category="success")
    return redirect(url_for('all_members'))


# shows all members
@app.route('/member/all/advanced_search')
@login_required
@requires_roles('admin')
def advanced_search():
    return render_template('advanced_search.html',
                           members=db.get_all_members(),
                           inactiveMembers=db.get_all_inactive_members(),
                           showInactive=False)


# Admin - Administration ###########

# shows all members
@app.route('/admin/all')
@login_required
@requires_roles('admin')
def all_admin():
    return render_template('admin_profiles.html',
                           admin=db.get_all_admin(),
                           inactiveAdmin=db.get_all_inactive_admin(),
                           showInactive=False)
