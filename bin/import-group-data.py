#!/usr/bin/env python3

import re
from openpyxl import load_workbook
from sys import argv


class HomeChurch(object):
    def __init__(self, id, covering, sector, name,
                 leader, leader_spouse,
                 assistant, assistant_spouse,
                 host, host_spouse):
        self.id = id
        self.covering = covering
        self.sector = sector
        self.name = name
        self.leader = leader
        self.leader_spouse = leader_spouse
        self.assistant = assistant
        self.assistant_spouse = assistant_spouse
        self.host = host
        self.host_spouse = host_spouse

    def __repr__(self):
        return "<HomeChurch {} {}>".format(self.id, self.name)

    def sql_insert(self):
        stmt = """INSERT INTO public.homegroup(id, name, location, description, is_active)
   VALUES(%(id)s, '%(name)s', '%(location)s', '%(sector)s', TRUE);"""
        return stmt % {'id': self.id,
                       'name': self.name,
                       'location': self.host.directions,
                       'sector': self.sector}


class Person(object):
    def __init__(self, first_name, last_name):
        self.first_name = first_name
        self.last_name = last_name


class Leader(Person):
    def __init__(self, first_name, last_name, home_phone, work_phone, mobile_phone, mobile_provider, email):
        super().__init__(first_name, last_name)
        self.home_phone = home_phone
        self.work_phone = work_phone
        self.mobile_phone = mobile_phone
        self.mobile_provider = mobile_provider
        self.email = email

    def __repr__(self):
        return "<Leader {} {} {}>".format(self.first_name, self.last_name, self.email)


class LeaderSpouse(Person):
    def __init__(self, first_name, last_name, mobile_phone, mobile_provider, email):
        super().__init__(first_name, last_name)
        self.mobile_phone = mobile_phone
        self.mobile_provider = mobile_provider
        self.email = email

    def __repr__(self):
        return "<LeaderSpouse {} {} {}>".format(self.first_name, self.last_name, self.email)


class Assistant(Person):
    def __init__(self, first_name, last_name, home_phone, mobile_phone, mobile_provider, email):
        super().__init__(first_name, last_name)
        self.home_phone = home_phone
        self.mobile_phone = mobile_phone
        self.mobile_provider = mobile_provider
        self.email = email

    def __repr__(self):
        return "<Assistant {} {} {}>".format(self.first_name, self.last_name, self.email)


class AssistantSpouse(Person):
    def __init__(self, first_name, last_name, email):
        super().__init__(first_name, last_name)
        self.email = email

    def __repr__(self):
        return "<AssistantSpouse {} {} {}>".format(self.first_name, self.last_name, self.email)


class Host(Person):
    def __init__(self, first_name, last_name, directions, home_phone, mobile_phone, mobile_provider):
        super().__init__(first_name, last_name)
        self.directions = directions
        self.home_phone = home_phone
        self.mobile_phone = mobile_phone
        self.mobile_provider = mobile_provider

    def __repr__(self):
        return "<Host {} {}>".format(self.first_name, self.last_name)


class HostSpouse(Person):
    def __init__(self, first_name, last_name):
        super().__init__(first_name, last_name)

    def __repr__(self):
        return "<HostSpouse {} {}>".format(self.first_name, self.last_name)


def is_empty_sheet(worksheet):
    return (worksheet.min_row == worksheet.max_row and
            worksheet.min_column == worksheet.max_column)


def val(worksheet, row_idx, col_idx):
    rtn = worksheet.cell(row_idx, col_idx).value
    if type(rtn) == 'string':
        rtn = rtn.strip()
    return rtn


def val_range(worksheet, row_idx, col_min, col_max):
    return [val(ws, row_idx, col_idx) for col_idx in range(col_min, col_max + 1)]


def parse_emails(cell_value):
    """Parse emails; assume first two are leader and spouse; ignore remainder."""
    rtn = ['', '']
    if cell_value is not None:
        emails = re.split(r'[; ]+', cell_value)
        if len(emails) >= 1:
            rtn[0] = emails[0]
        if len(emails) >= 2:
            rtn[1] = emails[1]
    return rtn


wbook = load_workbook(filename=argv[1], data_only=True, guess_types=True)

for ws in wbook:
    if is_empty_sheet(ws):
        continue
    print("Processing", ws.title)
    for row_idx in range(3, ws.max_row + 1):
        [leader_email, leader_spouse_email] = parse_emails(val(ws, row_idx, 14))
        leader = Leader(*val_range(ws, row_idx, 4, 9), leader_email)
        leader_spouse = LeaderSpouse(*val_range(ws, row_idx, 10, 13), leader_spouse_email)

        [assistant_email, assistant_spouse_email] = parse_emails(val(ws, row_idx, 22))
        assistant = Assistant(*val_range(ws, row_idx, 15, 19), assistant_email)
        assistant_spouse = AssistantSpouse(*val_range(ws, row_idx, 20, 21), assistant_spouse_email)

        host = Host(*val_range(ws, row_idx, 23, 28))
        host_spouse = HostSpouse(*val_range(ws, row_idx, 29, 30))

        # print(home_church,
        #       leader, leader_spouse,
        #       assistant, assistant_spouse,
        #       host, host_spouse)

        home_church = HomeChurch(*val_range(ws, row_idx, 1, 4),
                                 leader, leader_spouse,
                                 assistant, assistant_spouse,
                                 host, host_spouse)

        print(home_church.sql_insert())
